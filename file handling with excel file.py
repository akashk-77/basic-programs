
# Install openpyxl before running:


from openpyxl import Workbook, load_workbook


workbook = Workbook()

sheet = workbook.active


sheet.title = "StudentData"


sheet["A1"] = "ID"
sheet["B1"] = "Name"
sheet["C1"] = "Marks"


sheet["A2"] = 101
sheet["B2"] = "John"
sheet["C2"] = 85

sheet["A3"] = 102
sheet["B3"] = "Alice"
sheet["C3"] = 92

sheet["A4"] = 103
sheet["B4"] = "David"
sheet["C4"] = 78


workbook.save("students.xlsx")

print("Excel file created and data written successfully!")


read_workbook = load_workbook("students.xlsx")


read_sheet = read_workbook["StudentData"]

print("\nReading data from Excel file:\n")


for row in read_sheet.iter_rows(values_only=True):
    print(row)
